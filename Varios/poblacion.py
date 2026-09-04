import os
import re
import openpyxl
import pandas as pd
from io import StringIO
from dotenv import load_dotenv
from datetime import datetime
import psycopg2

# ==============================
# 1. Cargar variables desde .env
# ==============================
load_dotenv()

pg_user = os.getenv("PG_USER")
pg_pass = os.getenv("PG_PASS")
pg_host = os.getenv("PG_HOST")
pg_port = os.getenv("PG_PORT", "5433")
pg_db   = os.getenv("PG_DB")

EXCEL_PATH = r"C:\Users\gcpp.ggi.sge2\Desktop\GGI\poblacion\Centros Homologados_POBLACION_GCSPE_SGE.xlsx"
EXCEL_PATH_HOMOLOG = r"C:\Users\gcpp.ggi.sge2\Desktop\GGI\poblacion\homologacion_centros.xlsx"
CHUNK_SIZE = 50000

# ==============================
# 2. Utilidades
# ==============================
def normalizar_columna(nombre):
    nombre = nombre.strip().lower()
    nombre = nombre.replace("-", "_").replace(" ", "_")
    nombre = re.sub(r"[^a-z0-9_]", "", nombre)
    return nombre


def limpiar_df(df):
    for col in df.select_dtypes(include=["object"]).columns:
        df[col] = (
            df[col]
            .astype(str)
            .replace({"None": ""})
            .str.replace("\x00", "", regex=False)
            .str.encode("utf-8", "ignore")
            .str.decode("utf-8")
        )
    return df


def limpiar_rango_edad(df):
    # "100 A MAS" -> "100"; el resto de valores ya vienen como el número puro.
    df["rango_edad"] = df["rango_edad"].str.extract(r"(\d+)")
    return df


def copiar_a_postgres(cursor, df, table_name, columnas):
    csv_buffer = StringIO()
    df.to_csv(csv_buffer, index=False, header=False)
    csv_buffer.seek(0)
    cursor.copy_expert(
        sql=f"COPY {table_name} ({', '.join(columnas)}) FROM STDIN WITH CSV",
        file=csv_buffer,
    )


# ==============================
# 3. Conexión PostgreSQL destino
# ==============================
print(f"Conectando a PostgreSQL en {pg_host}:{pg_port}, base de datos: {pg_db}...")
conn_pg = psycopg2.connect(
    host=pg_host,
    port=pg_port,
    dbname=pg_db,
    user=pg_user,
    password=pg_pass,
)
conn_pg.autocommit = False
cursor_pg = conn_pg.cursor()
print("Conexión a PostgreSQL establecida.")

start_time = datetime.now()
print(f"\n Inicio del ETL: {start_time:%Y-%m-%d %H:%M:%S}")

try:
    # ==============================
    # 4. Hoja "Base" -> dssge.poblacion_gcspe_sge
    # ==============================
    tabla_base = "dssge.poblacion_gcspe_sge"
    print(f"\n--- Procesando hoja 'Base' -> {tabla_base} ---")

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws_base = wb["Base"]

    filas_iter = ws_base.iter_rows(values_only=True)
    columnas_originales = list(next(filas_iter))
    columnas_base = [normalizar_columna(c) for c in columnas_originales]

    print(f"Columnas detectadas: {columnas_base}")

    # La tabla destino ya existía como placeholder sin columnas; se recrea con la estructura correcta.
    cursor_pg.execute(f"DROP TABLE IF EXISTS {tabla_base};")
    cursor_pg.execute(f"""
        CREATE TABLE {tabla_base} (
            poblacion_acumulada       VARCHAR(6),
            ubigeo_inei               VARCHAR(10),
            departamento              VARCHAR(100),
            provincia                 VARCHAR(100),
            distrito                  VARCHAR(100),
            cod_centro                VARCHAR(10),
            nombre_centro_asistencial VARCHAR(200),
            nivel_centro_asistencial  VARCHAR(50),
            tipo_centro_asistencial   VARCHAR(100),
            nombre_red                VARCHAR(100),
            condicion                 VARCHAR(50),
            parentesco                VARCHAR(50),
            tipo_asegurado            VARCHAR(50),
            tipo_seguro               VARCHAR(50),
            tipo_atencion             VARCHAR(100),
            sexo                      VARCHAR(20),
            rango_edad                VARCHAR(20),
            cantidad_asegurados       INTEGER
        );
    """)
    print(f"Tabla {tabla_base} recreada con estructura correcta.")

    lote = []
    filas_procesadas = 0
    num_lote = 1
    for fila in filas_iter:
        lote.append(fila)
        if len(lote) >= CHUNK_SIZE:
            df_lote = pd.DataFrame(lote, columns=columnas_base)
            df_lote = limpiar_df(df_lote)
            df_lote = limpiar_rango_edad(df_lote)
            copiar_a_postgres(cursor_pg, df_lote, tabla_base, columnas_base)
            filas_procesadas += len(lote)
            print(f"  Lote {num_lote}: {len(lote)} filas cargadas (acumulado {filas_procesadas}).")
            lote = []
            num_lote += 1

    if lote:
        df_lote = pd.DataFrame(lote, columns=columnas_base)
        df_lote = limpiar_df(df_lote)
        df_lote = limpiar_rango_edad(df_lote)
        copiar_a_postgres(cursor_pg, df_lote, tabla_base, columnas_base)
        filas_procesadas += len(lote)
        print(f"  Lote {num_lote} (final): {len(lote)} filas cargadas (acumulado {filas_procesadas}).")

    wb.close()
    print(f"Hoja 'Base' cargada: {filas_procesadas} filas en total.")

    # ==============================
    # 5. homologacion_centros.xlsx -> dssge.poblacion_gcspe_sge_homologacion
    # ==============================
    tabla_homolog = "dssge.poblacion_gcspe_sge_homologacion"
    print(f"\n--- Procesando {EXCEL_PATH_HOMOLOG} -> {tabla_homolog} ---")

    wb2 = openpyxl.load_workbook(EXCEL_PATH_HOMOLOG, read_only=True, data_only=True)
    ws_homolog = wb2[wb2.sheetnames[0]]

    filas_homolog = list(ws_homolog.iter_rows(values_only=True))
    columnas_homolog = [normalizar_columna(c) for c in filas_homolog[0]]
    datos_homolog = filas_homolog[1:]
    wb2.close()

    print(f"Columnas detectadas: {columnas_homolog}")

    cursor_pg.execute(f"DROP TABLE IF EXISTS {tabla_homolog};")
    cursor_pg.execute(f"""
        CREATE TABLE {tabla_homolog} (
            cod_centro_essi_gcspe            VARCHAR(10),
            nombre_red_gcspe                 VARCHAR(100),
            nombre_centro_asistencial_gcspe  VARCHAR(200),
            tipo_sge                         VARCHAR(50),
            tipo_sge1                        VARCHAR(50),
            sge1_denom_larga_sge             VARCHAR(200),
            sge1_denom_corta_sge             VARCHAR(100),
            observacion                      VARCHAR(500)
        );
    """)
    print(f"Tabla {tabla_homolog} recreada con estructura correcta.")

    df_homolog = pd.DataFrame(datos_homolog, columns=columnas_homolog)
    df_homolog = limpiar_df(df_homolog)
    copiar_a_postgres(cursor_pg, df_homolog, tabla_homolog, columnas_homolog)
    print(f"Hoja 'homologacion' cargada: {len(df_homolog)} filas en total.")

    # ==============================
    # 6. Commit
    # ==============================
    conn_pg.commit()
    print("\nCambios confirmados (commit) en PostgreSQL.")

except Exception as e:
    conn_pg.rollback()
    print(f"\nError durante el ETL, se revirtieron los cambios: {e}")
    raise

finally:
    cursor_pg.close()
    conn_pg.close()
    print("Conexión a PostgreSQL cerrada.")

end_time = datetime.now()
print(f"\n Fin del ETL: {end_time:%Y-%m-%d %H:%M:%S}")
print(f" Duración total: {end_time - start_time}")
